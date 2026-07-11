from openpyxl import load_workbook

from app import app
from extensions import db
from models.object import YLFFObject


EXCEL_FILE = "/app/YLFF-TERITORIJAS.xlsx"


def parse_coordinates(value):
    value = str(value).strip()

    value = value.replace("N", " N")
    value = value.replace("S", " S")
    value = value.replace("E", " E")
    value = value.replace("W", " W")

    parts = value.split()

    if len(parts) < 4:
        print(f"BAD COORDINATES: {value}")
        return None, None

    lat = float(parts[0])

    if parts[1] == "S":
        lat = -lat

    lon = float(parts[2])

    if parts[3] == "W":
        lon = -lon

    return lat, lon


def format_reference(reference):
    reference = str(reference).strip().upper()

    if "-" not in reference:
        return reference

    prefix, number = reference.split("-", 1)

    number = number.zfill(4)

    return f"{prefix}-{number}"


with app.app_context():
    workbook = load_workbook(EXCEL_FILE)

    sheet = workbook.active

    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue

        reference = format_reference(row[0])

        name = str(row[1]).strip()

        coordinates = row[2]

        locator = None

        if len(row) > 3 and row[3]:
            locator = str(row[3]).strip()

        latitude = None
        longitude = None

        if coordinates:
            latitude, longitude = parse_coordinates(
                coordinates
            )

        item = YLFFObject.query.filter_by(
            reference=reference
        ).first()

        if item:
            item.name = name
            item.latitude = latitude
            item.longitude = longitude
            item.locator = locator

            print(f"UPDATE {reference}")

            continue

        item = YLFFObject(
            reference=reference,
            name=name,
            latitude=latitude,
            longitude=longitude,
            locator=locator,
            is_active=True,
            status="active",
        )

        db.session.add(item)

        print(f"ADD {reference}")

    db.session.commit()

    print("IMPORT FINISHED")